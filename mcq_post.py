from openpyxl.workbook import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
import os

def insertImg(ws,idImgPath, pre_rows, col):

    img = Image(idImgPath)
    ###########################
    row_cut = 8000
    col_cut = 12500
    
    _from = AnchorMarker(
        col=col,
        row=pre_rows-1,
        colOff=col_cut,
        rowOff=row_cut,
    )
    to = AnchorMarker(
        col=col + 1,
        row=pre_rows,
        colOff=-col_cut,
        rowOff=-row_cut,
    )
    img.anchor = TwoCellAnchor(editAs="twoCell", _from=_from, to=to)
    ws.add_image(img)

    return ws
def post_processing(val, save_path, fileNames, tempPath, index):
    
    if len(val) > 0:
        wb = Workbook()
        ws = wb.active
        ws.title = "new table"        
        pre_rows = 1
        ws.cell(row=pre_rows, column=1).value = "SHEET NUMBER"
        ws.cell(row=pre_rows, column=1).font = Font(bold=True)        
        ws.cell(row=pre_rows, column=2).value = "CIRCLE SEAT NUMBER"
        ws.cell(row=pre_rows, column=2).font = Font(bold=True)
        ws.cell(row=pre_rows, column=3).value = "HANDWRITTEN SEAT NUMBER"
        ws.cell(row=pre_rows, column=3).font = Font(bold=True)        
        ws.cell(row=pre_rows, column=4).value = "CIRCLE BOOKLET VERSION NUMBER"
        ws.cell(row=pre_rows, column=4).font = Font(bold=True)
        ws.cell(row=pre_rows, column=5).value = "HANDWRITTEN BOOKLET VERSION NUMBER"
        ws.cell(row=pre_rows, column=5).font = Font(bold=True)        
        ws.cell(row=pre_rows, column=26).value = "SHEET NUMBER IMAGE"
        ws.cell(row=pre_rows, column=26).font = Font(bold=True)      
        ws.cell(row=pre_rows, column=27).value = "SHEET CROP PATH"
        ws.cell(row=pre_rows, column=27).font = Font(bold=True)      
        ws.cell(row=pre_rows, column=28).value = "SEAT NUMBER IMAGE"
        ws.cell(row=pre_rows, column=28).font = Font(bold=True) 
        ws.cell(row=pre_rows, column=29).value = "SEAT CROP PATH"
        ws.cell(row=pre_rows, column=29).font = Font(bold=True)
        ws.cell(row=pre_rows, column=30).value = "BOOKLET IMAGE"
        ws.cell(row=pre_rows, column=30).font = Font(bold=True)  
        ws.cell(row=pre_rows, column=31).value = "BOOKLET CROP PATH"
        ws.cell(row=pre_rows, column=31).font = Font(bold=True)  
        ws.cell(row=pre_rows, column=32).value = "FILE NAME"
        ws.cell(row=pre_rows, column=32).font = Font(bold=True)  

        for i in range(20):
            ws.cell(row=pre_rows, column=6+i).value = f"Q{i+1}"
            ws.cell(row=pre_rows, column=6+i).font = Font(bold=True)
        pre_rows = pre_rows + 1
        thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')) 
        ### loopping every document in multi document ###

        for ii, fileName in enumerate(fileNames):   
        # for v in val:
            nm = fileName.split('/')[-1]
            sheetImgPath = f"{tempPath}/{index[0]}_{nm}" 
            idImgPath = f"{tempPath}/{index[1]}_{nm}" 
            queImgPath = f"{tempPath}/{index[2]}_{nm}" 
            v = val[ii]  
            try:
                sheet_num, seat_num, seat_num_hand, question, question_hand, answer = v
                if seat_num is not None: ws.cell(row=pre_rows, column=1).value = sheet_num
                else: ws.cell(row=pre_rows, column=1).value = 'Not_Reconized'
                if seat_num is not None: ws.cell(row=pre_rows, column=2).value = seat_num
                else: ws.cell(row=pre_rows, column=2).value = 'Not_Reconized'
                if seat_num_hand is not None: ws.cell(row=pre_rows, column=3).value = seat_num_hand
                else: ws.cell(row=pre_rows, column=3).value = 'Not_Reconized'                
                if question is not None: ws.cell(row=pre_rows, column=4).value = question
                else: ws.cell(row=pre_rows, column=4).value = 'Not_Reconized'
                if question_hand is not None: ws.cell(row=pre_rows, column=5).value = question_hand
                else: ws.cell(row=pre_rows, column=5).value = 'Not_Reconized'                
                for i in range(len(answer)):
                    if answer[i] is not None:
                        ws.cell(row=pre_rows, column=6+i).value = answer[i]
                    else:
                        ws.cell(row=pre_rows, column=6+i).value = 'Not_Reconized'
                
                ws.cell(row=pre_rows, column=27).value = sheetImgPath # SHEET CROP PATH
                ws.cell(row=pre_rows, column=29).value = idImgPath # SEAT CROP PATH
                ws.cell(row=pre_rows, column=31).value = queImgPath # VERSION CROP PATH

                insertImg(ws, sheetImgPath, pre_rows, 25)
                insertImg(ws, idImgPath, pre_rows, 27)
                insertImg(ws, queImgPath, pre_rows, 29)
                ws.cell(row=pre_rows, column=32).value = fileName
                pre_rows = pre_rows+1
            except: pass

        # cell swrap, thin
        row_no = 1
        for i in ws.rows:
            for j in range(len(i)):
                ws[get_column_letter(j+1)+str(row_no)].alignment = Alignment(wrap_text=True, vertical='center',horizontal='center')
            row_no = row_no + 1   
        for i in range(1, pre_rows):       
            for j in range(1, 33):
                ws.cell(row=i, column=j).border = thin_border
        for i in range(27):
            if i < 5 or i>24:
                ws.column_dimensions[get_column_letter(i+1)].width = 16
            else:
                ws.column_dimensions[get_column_letter(i+1)].width = 5
        ws.column_dimensions[get_column_letter(26)].width = 40
        ws.column_dimensions[get_column_letter(27)].width = 40
        ws.column_dimensions[get_column_letter(28)].width = 40
        ws.column_dimensions[get_column_letter(29)].width = 40
        ws.column_dimensions[get_column_letter(30)].width = 40
        ws.column_dimensions[get_column_letter(31)].width = 40
        ws.column_dimensions[get_column_letter(32)].width = 40
        wb.save(save_path)
    else:
        print("=== Table of this pdf is not detected ===")

    return None
