from PyQt5 import QtWidgets, uic,QtGui, QtCore
from PyQt5.QtWidgets import QMainWindow, QApplication, QVBoxLayout, QDialog
from PyQt5.QtCore import pyqtSignal, pyqtSlot
import sys, os, getpass
from main_mcq import mcq
from threading import Thread
import uuid, shutil, json
from AESCipher import AESCipher
import pandas
import openpyxl
#uuid.getnode()
class MainPage(QMainWindow):
    def __init__(self):
        super(MainPage,self).__init__()
        uic.loadUi(resource_path("main.ui"),self)
        self.cnt2=0
        self.totol2=0
        self.cnt1=0
        self.totol1=0
        self.templatePath = os.path.join(logPath,'template')
        self.initUI() 
    
    def initUI(self):
        self.btn_browse1.clicked.connect(self.btnSelectFile1)
        self.btn_process1.clicked.connect(self.btnProcessFile1)
        self.pbar_tab1.valueChanged.connect(self.onChangeValue1)
        self.single_done1.connect(self.progress1)
        self.lnk_tab1.clicked.connect(self.onClickLblTab1)
        self.lnk_tab1.setEnabled(False)
        # self.selectStyles = self.getConfig()
    
    def getConfig(self):
        configFiles = os.listdir(self.configPath)
        return [v.split('.')[0] for v in configFiles if 'style' in v]

    def btnSelectFile1(self):
        self.configPath, _ = QtWidgets.QFileDialog.getOpenFileNames(
            self, "Open File", "", "CSV Files(*.json);;All Files (*)",
            options=QtWidgets.QFileDialog.DontUseNativeDialog)
        try:
            self.configPath = self.configPath[0]
            self.le_tab1.setText(self.configPath)   
        except: pass
    
    def btnProcessFile1(self):
        self.btn_process1.setEnabled(False)
        t1=Thread(target=self.Operation1)
        t1.start()

    def Operation1(self):
        # mcq(self, self.dialog)
        
        try:
            self.configPath = self.le_tab1.text()
            with open(self.configPath, "r") as f:
                jsonInfos = json.load(f)
            sourcePath = jsonInfos['excelFilePath']
            self.outputPath = jsonInfos['outputPath']
            makedir(self.outputPath)
            # try: clear_contents(outputPath)
            # except: pass            
        except: 
            self.lbl1_tab1.setText("Wanring: Please Select Config File Exactly")
            self.btn_process1.setEnabled(True)
            return None
        # wb_obj = openpyxl.load_workbook(sourcePath)
        # sheet_obj = wb_obj.active
        source = pandas.read_excel(sourcePath).to_numpy()
        batchName, cropPathBody = None, "CROP"
        img_list = []
        self.lbl1_tab1.setText("MCQ scan starting ...")        
        ind = []
        for i, item in enumerate(source):
            if str(item[2]) == 'None' or str(item[2]) == 'nan': break
            if item[2].lower() == 'n':
                if batchName != item[0]:
                    try: 
                        self.process(out, img_list, os.path.join(out, cropPathBody) , batchName, ind, sourcePath)
                        img_list = []
                    except: pass
                    batchName = item[0]
                    out = os.path.join(self.outputPath, batchName)
                img_list.append(item[1])
                ind.append(i)
        if len(img_list) > 0:
            result = self.process(out, img_list, os.path.join(out, cropPathBody), batchName, ind, sourcePath)
        
        # if result: 
        #     self.lnk_tab1.setEnabled(True)
        #     self.lnk_tab1.setText("Go Output Folder")
        #     self.lbl1_tab1.setText("Successfuly done") 
            
        # else: self.lbl1_tab1.setText("Failed") 
        self.lnk_tab1.setEnabled(True)
        self.lnk_tab1.setText("Go Output Folder")
        self.lbl1_tab1.setText("Done!") 
        self.btn_process1.setEnabled(True)
       
    def process(self, out, img_list, cropPathBody, batchName, ind, sourcePath):
    
        makedir(out)
        # vals, labels = barcode(img_list, out, cropPathBody) # 1 means file name, and 0 means qrcode
        save_path = out + f"/MCQ_{batchName}.xlsx"    
        ID_index = ['sheet', 'seat', 'version']    
        vals, fileNames = mcq(self, img_list, out, cropPathBody, save_path, ID_index)
        tempSourcePath = '/'.join(sourcePath.split('/')[0:-1] + ['temp.xlsx'])
               
        if vals:
            try: shutil.copyfile(sourcePath, tempSourcePath)
            except: pass
            wb = openpyxl.load_workbook(tempSourcePath)
            ws = wb.active   
                           
            # json output save
            jsonOut = {}
            for i, val in enumerate(vals):
                name = fileNames[i].split('/')[-1]
                value = {}
                value["BATCHNAME"] = batchName
                value["SHEET NUMBER"] = val[0]
                value["CIRCLE SEAT NUMBER"] = val[1]
                value["HANDWRITTEN SEAT NUMBER"] = val[2]
                value["CIRCLE BOOKLET VERSION NUMBER"] = val[3]
                value["HANDWRITTEN BOOKLET VERSION NUMBER"] = val[4]
                value["ANSWER"] = ''.join(val[5])
                value["SHEET CROP PATH"] = cropPathBody.replace("\\", "/") + f"/{ID_index[0]}_{name}"
                value["SEAT CROP PATH"] = cropPathBody.replace("\\", "/") + f"/{ID_index[1]}_{name}"
                value["VERSION CROP PATH"] = cropPathBody.replace("\\", "/") + f"/{ID_index[2]}_{name}"
                jsonOut[fileNames[i]] = value
            with open(os.path.join(out, f"MCQ_{batchName}.json"), 'w') as f:
                json.dump(jsonOut, f, indent=4, sort_keys=True)    
            for i in ind:
                ws.cell(i+2,3).value = 'Y' 
            
            wb.save(sourcePath)
            try: os.remove(tempSourcePath)
            except: pass
            
            dst_dir = os.path.join(out, 'input') 
            makedir(dst_dir)
            for imd in img_list:
                try: 
                    imgs = [f for f in os.listdir(imd) if (f.split('.')[-1].lower() in ['png','jpg', 'tif'])]
                    for img in imgs:
                        shutil.copy(os.path.join(imd, img), os.path.join(dst_dir, img))  
                except: pass              
        
            return True
        else:
            try: shutil.copyfile(sourcePath, tempSourcePath)
            except: pass
            wb = openpyxl.load_workbook(tempSourcePath)
            ws = wb.active               
            for i in ind:
                ws.cell(i+1,3).value = 'N' 
            wb.save(sourcePath)
            return False      
    def onChangeValue1(self,val):
        self.pbar_tab1.setFormat(str(self.cnt1) + '/' + str(self.total1))

    single_done1 = pyqtSignal()
    @pyqtSlot()
    def progress1(self):
        self.pbar_tab1.setValue(int((self.cnt1/self.total1)*100))

    def openFolder1(self, path):
        # self.lbl1_tab1.setText("Result : Successfully processed ")
        self.path1=path
        self.lnk_tab1.setText(str(path))
        self.openOutput(path)
    
    def onClickLblTab1(self):
        # self.openOutput(self.path1)
        self.openFolder1(self.outputPath)
    
    def openOutput(self,path): 
        if path:  
            print(path, "path")
            QtGui.QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile(str(path)))

class KeyWindow(QMainWindow):
    def __init__(self):
        super(KeyWindow,self).__init__()
        self.node = str(uuid.getnode())
        uic.loadUi(resource_path("keyWindow.ui"),self)
        self.lbl_id.setText(self.node)
        self.initUI()

    def initUI(self):
        self.btn_submit.clicked.connect(self.onSubmit)
    
    def onSubmit(self):
        
        if self.txt_key.toPlainText():
            fp = open(logPath+'/.validate', 'w')
            fp.write(self.txt_key.toPlainText())
            fp.close()

            if validate():
                self.win = MainPage()
                self.win.show()
                self.hide()
            else:
                self.lbl_msg.setText("Invalid key")
def window():
    app = QApplication(sys.argv)
    win = MainPage()
    win.show()
    sys.exit(app.exec_())

def windowValidate():
    app = QApplication(sys.argv)
    win = KeyWindow()
    win.show()
    sys.exit(app.exec_())

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def validate():
    try:
        fp =open(logPath+'/.validate','rb')
        data = fp.read()
        fp.close()
    except: return False
    if data:
        try:
            c = AESCipher()
            if str(uuid.getnode()) == c.decrypt(data):
                return True
        except:
            pass
    return False
def makedir(dir):
    try:
        os.mkdir(dir)
    except:
        pass
def clear_contents(dir_path):
    '''
    Deletes the contents of the given filepath. Useful for testing runs.
    '''
    filelist = os.listdir(dir_path)
    if filelist:
        for f in filelist:
            if os.path.isdir(os.path.join(dir_path, f)):
                shutil.rmtree(os.path.join(dir_path, f))
            else:
                os.remove(os.path.join(dir_path, f))
    return None
logPath = f"C:/Users/{getpass.getuser()}/.Mcq"
makedir(logPath)  
try: shutil.copytree("template", os.path.join(logPath, 'template'))
except: pass
if not validate():
    windowValidate()
else:
    window()
